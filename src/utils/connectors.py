"""
URL connectors — each function accepts a file path and returns a ``list[str]``
of URLs extracted from that file.

Dispatcher
----------
read_file(path)     Auto-selects the right connector based on the file extension.

Individual connectors
---------------------
_read_txt(path)      Plain-text file, one URL per line; blank lines and lines
                    starting with '#' are ignored.
_read_json(path)     JSON array of objects, each expected to have a "url" key
                    (case-insensitive). Plain string items are also accepted.
_read_xlsx(path)     Excel workbook (.xlsx); URLs are read from a column whose
                    header matches "URL" (case-insensitive) in the active sheet.
_read_csv(path)      CSV file with a header row; URLs are read from a column
                    matching ``url_column`` (case-insensitive, default ``"url"``).
"""

import csv
import json
import logging
from pathlib import Path

import openpyxl
from tqdm import tqdm
from tqdm.contrib.logging import logging_redirect_tqdm

logger = logging.getLogger(__name__)


def _read_txt(path: str) -> list[str]:
    """Read URLs from a plain-text file, one per line.

    Blank lines and lines whose first non-whitespace character is '#' are
    silently skipped.

    Example file::

        # Production pages
        https://example.com
        https://example.com/about
    """
    logger.debug("Opening TXT file: '%s'.", path)
    with open(path) as f:
        urls = [
            line.strip()
            for line in f
            if line.strip() and not line.strip().startswith("#")
        ]
    logger.info("Read %d URL(s) from '%s'.", len(urls), path)
    return urls


def _read_json(path: str) -> list[str]:
    """Read URLs from a JSON file containing a top-level array.

    Each item in the array may be:

    * A plain string — used as-is.
    * An object — the value of the first key whose name is ``"url"``
      (case-insensitive) is used.

    Raises ``ValueError`` if the top-level value is not an array, or if an
    object item contains no ``"url"`` key.

    Example file::

        [
          {"name": "Home",  "url": "https://www.example.com"},
          {"name": "About", "url": "https://www.example.com/about"}
        ]
    """
    logger.debug("Opening JSON file: '%s'.", path)
    with open(path) as f:
        data = json.load(f)

    if not isinstance(data, list):
        raise ValueError(
            f"Expected a JSON array at the top level, got {type(data).__name__!r}."
        )

    logger.debug("Loaded JSON array with %d item(s).", len(data))
    urls: list[str] = []
    for i, item in enumerate(data):
        if isinstance(item, str):
            if item.strip():
                urls.append(item.strip())
            else:
                logger.debug("Skipping empty string at index %d.", i)
        elif isinstance(item, dict):
            url = next(
                (v for k, v in item.items() if k.lower() == "url"),
                None,
            )
            if url is None:
                raise ValueError(
                    f"Item at index {i} has no 'url' key. Keys present: "
                    f"{list(item.keys())!r}"
                )
            urls.append(str(url).strip())
        else:
            raise ValueError(
                f"Item at index {i} must be a string or object, "
                f"got {type(item).__name__!r}: {item!r}"
            )

    logger.info("Read %d URL(s) from '%s'.", len(urls), path)
    return urls


def _read_xlsx(
    path: str, sheet_name: str | None = None, url_column: str = "URL"
) -> list[str]:
    """Read URLs from an Excel workbook (.xlsx).

    If ``sheet_name`` is provided, that sheet is used; otherwise the workbook's
    active sheet is used. The first row is treated as a header row and a column
    matching ``url_column`` (case-insensitive, default ``"URL"``) must be
    present. All subsequent non-empty cells in that column are returned as URLs.

    Raises ``ValueError`` if the sheet name is not found, the sheet is empty,
    or no column matching ``url_column`` is present.

    Example sheet::

        | Name  | URL                          | Description |
        |-------|------------------------------|-------------|
        | Home  | https://www.example.com      | ...         |
        | About | https://www.example.com/about| ...         |
    """
    logger.debug("Opening Excel workbook: '%s'.", path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name is not None:
            if sheet_name not in wb.sheetnames:
                raise ValueError(
                    f"Sheet {sheet_name!r} not found in the Excel file. "
                    f"Available sheets: {wb.sheetnames!r}"
                )
            ws = wb[sheet_name]
            logger.info("Using sheet: '%s' (specified).", sheet_name)
        else:
            ws = wb.active
            if ws is None:
                raise ValueError("The Excel file has no active sheet.")
            logger.info("Using sheet: '%s' (active).", ws.title)
        if ws is None:
            raise ValueError("The Excel file has no active sheet.")
        row_iter = ws.iter_rows(values_only=True)

        header = next(row_iter, None)
        if header is None:
            raise ValueError("The Excel file appears to be empty.")

        url_col_idx = next(
            (
                i
                for i, h in enumerate(header)
                if h and str(h).strip().upper() == url_column.strip().upper()
            ),
            None,
        )
        if url_col_idx is None:
            found = [str(h) for h in header if h is not None]
            raise ValueError(
                f"No {url_column!r} column found in the Excel file. Headers found: {found!r}"
            )
        logger.debug(
            "Found %r column at index %d (header: '%s').",
            url_column,
            url_col_idx,
            header[url_col_idx],
        )

        total_rows = (ws.max_row or 1) - 1  # subtract header row
        urls: list[str] = []
        with logging_redirect_tqdm():
            for row in tqdm(
                row_iter,
                total=total_rows,
                desc=f"Reading '{ws.title}'",
                unit="row",
            ):
                if url_col_idx < len(row):
                    cell = row[url_col_idx]
                    if cell is not None:
                        value = str(cell).strip()
                        if value:
                            logger.debug("Found URL: '%s'.", value)
                            urls.append(value)
    finally:
        wb.close()

    logger.info("Read %d URL(s) from sheet '%s' in '%s'.", len(urls), ws.title, path)
    return urls


def _read_csv(path: str, url_column: str = "url") -> list[str]:
    """Read URLs from a CSV file with a header row.

    The first row is treated as a header row. A column matching ``url_column``
    (case-insensitive, default ``"url"``) must be present. All subsequent
    non-empty cells in that column are returned as URLs.

    Raises ``ValueError`` if the file is empty or no column matching
    ``url_column`` is found.

    Example file::

        name,url
        Home,https://www.example.com
        About,https://www.example.com/about
    """
    logger.debug("Opening CSV file: '%s'.", path)
    urls: list[str] = []
    with open(path, newline="") as f:
        reader = csv.DictReader(f)

        if not reader.fieldnames:
            raise ValueError("The CSV file appears to be empty.")

        matched_col = next(
            (
                col
                for col in reader.fieldnames
                if col.strip().lower() == url_column.strip().lower()
            ),
            None,
        )
        if matched_col is None:
            raise ValueError(
                f"No {url_column!r} column found in the CSV file. "
                f"Headers found: {list(reader.fieldnames)!r}"
            )
        logger.debug("Found %r column (header: '%s').", url_column, matched_col)

        for row in reader:
            value = (row[matched_col] or "").strip()
            if value:
                logger.debug("Found URL: '%s'.", value)
                urls.append(value)

    logger.info("Read %d URL(s) from '%s'.", len(urls), path)
    return urls


# ---------------------------------------------------------------------------
# Dispatcher
# ---------------------------------------------------------------------------

_CONNECTORS: dict[str, callable] = {  # type: ignore[type-arg]
    ".txt": _read_txt,
    ".json": _read_json,
    ".xlsx": _read_xlsx,
    ".csv": _read_csv,
}


def read_file(
    path: str, sheet_name: str | None = None, url_column: str = "URL"
) -> list[str]:
    """Dispatch to the appropriate connector based on the file's extension.

    Supported extensions: ``.txt``, ``.json``, ``.xlsx``, ``.csv``

    ``sheet_name`` is only applicable to ``.xlsx`` files; it selects the sheet
    to read from. When ``None``, the workbook's active sheet is used.

    ``url_column`` applies to ``.xlsx`` and ``.csv`` files; it specifies the
    header name of the column containing URLs (case-insensitive). Defaults to
    ``"URL"`` for ``.xlsx`` and ``"url"`` for ``.csv``.

    Raises ``ValueError`` for unrecognised extensions.
    """
    ext = Path(path).suffix.lower()
    connector = _CONNECTORS.get(ext)
    if connector is None:
        supported = ", ".join(_CONNECTORS)
        raise ValueError(
            f"Unsupported file type {ext!r}. Supported extensions: {supported}"
        )
    logger.debug("Detected file type '%s', dispatching to %s.", ext, connector.__name__)
    if ext == ".xlsx":
        return connector(path, sheet_name=sheet_name, url_column=url_column)
    if ext == ".csv":
        return connector(path, url_column=url_column)
    return connector(path)
